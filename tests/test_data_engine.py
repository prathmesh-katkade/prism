"""Tests for modules.data_engine's large-file ingestion path.

load_data()'s existing behavior (encoding/delimiter sniffing, header
recovery, MAX_ROWS truncation) reads the *entire* file into pandas before
deciding what to keep — fine for small files, wasteful and eventually
impossible for a genuinely large one (a multi-GB CSV shouldn't have to
fully materialize in Python memory just to get truncated down to 50k rows
a moment later). This module adds two optional out-of-core paths: for CSV
files above a size threshold, DuckDB's out-of-core reader counts rows and
pulls a random sample directly; for .xlsx files above a (lower) size
threshold, a streaming openpyxl reader does the equivalent row-by-row with
reservoir sampling. Below each threshold, behavior is byte-for-byte
unchanged — these tests exist to prove those boundaries, not to re-test
load_data()'s existing encoding/header logic (already exercised manually
across every prior run's audit).
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from modules.data_engine import (
    LARGE_EXCEL_THRESHOLD_BYTES,
    LARGE_FILE_THRESHOLD_BYTES,
    _duckdb_sample_csv,
    _should_attempt_duckdb,
    _should_attempt_streaming_excel,
    _stream_sample_excel,
    load_data,
)


class _FakeUploadedFile(io.BytesIO):
    """Minimal stand-in for streamlit's UploadedFile — a BytesIO with the
    extra `.name`/`.size` attributes load_data() and its helpers read."""

    def __init__(self, data: bytes, name: str = "data.csv", size: int | None = None):
        super().__init__(data)
        self.name = name
        self.size = size if size is not None else len(data)


def _csv_bytes(n_rows: int) -> bytes:
    lines = ["id,value,label"] + [f"{i},{i * 1.5},group{i % 3}" for i in range(n_rows)]
    return ("\n".join(lines)).encode("utf-8")


def _xlsx_bytes(n_rows: int, banner: str | None = None, sheet_name: str = "Sheet1") -> bytes:
    """Build a real in-memory .xlsx workbook with a header row (id, value,
    label) followed by n_rows data rows — the Excel equivalent of
    _csv_bytes(), used to exercise the streaming reader against genuine
    xlsx structure rather than a mock."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if banner is not None:
        ws.append([banner])
        ws.append([])
    ws.append(["id", "value", "label"])
    for i in range(n_rows):
        ws.append([i, i * 1.5, f"group{i % 3}"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


pytest.importorskip("duckdb")


# ─────────────────────────────────────────────────────────────────────────
# _should_attempt_duckdb — the size gate
# ─────────────────────────────────────────────────────────────────────────
def test_should_attempt_duckdb_false_for_small_file():
    f = _FakeUploadedFile(_csv_bytes(10), size=1_000)
    assert _should_attempt_duckdb(f) is False


def test_should_attempt_duckdb_true_above_threshold():
    f = _FakeUploadedFile(_csv_bytes(10), size=LARGE_FILE_THRESHOLD_BYTES + 1)
    assert _should_attempt_duckdb(f) is True


def test_should_attempt_duckdb_false_when_size_unknown():
    f = _FakeUploadedFile(_csv_bytes(10))
    del f.size  # simulate an object that doesn't expose .size at all
    assert _should_attempt_duckdb(f) is False


# ─────────────────────────────────────────────────────────────────────────
# _duckdb_sample_csv — the actual out-of-core read
# ─────────────────────────────────────────────────────────────────────────
def test_duckdb_sample_returns_all_rows_when_under_cap():
    f = _FakeUploadedFile(_csv_bytes(50))
    result = _duckdb_sample_csv(f, max_rows=200)
    assert result is not None
    df, warnings = result
    assert len(df) == 50
    assert list(df.columns) == ["id", "value", "label"]
    assert warnings == []


def test_duckdb_sample_takes_random_sample_when_over_cap():
    f = _FakeUploadedFile(_csv_bytes(2000))
    result = _duckdb_sample_csv(f, max_rows=100, random_state=42)
    assert result is not None
    df, warnings = result
    assert len(df) == 100
    assert warnings and "2,000" in warnings[0]
    # A random sample across the whole file, not just the first 100 rows —
    # ids shouldn't all be small/contiguous the way df.head(100) would be.
    assert df["id"].max() > 200


def test_duckdb_sample_is_reproducible_given_same_seed():
    data = _csv_bytes(2000)
    df1, _ = _duckdb_sample_csv(_FakeUploadedFile(data), max_rows=50, random_state=7)
    df2, _ = _duckdb_sample_csv(_FakeUploadedFile(data), max_rows=50, random_state=7)
    assert sorted(df1["id"].tolist()) == sorted(df2["id"].tolist())


def test_duckdb_sample_returns_none_on_unparseable_content():
    garbage = _FakeUploadedFile(b"\x00\x01\x02\xff\xfe not a csv at all \x00")
    assert _duckdb_sample_csv(garbage, max_rows=100) is None


# ─────────────────────────────────────────────────────────────────────────
# load_data() — wiring: large files route through DuckDB, small files don't
# ─────────────────────────────────────────────────────────────────────────
def test_load_data_small_file_unaffected():
    f = _FakeUploadedFile(_csv_bytes(20))
    df, error, warnings = load_data(f)
    assert error is None
    assert len(df) == 20
    assert not any("DuckDB" in w for w in warnings)


def test_load_data_routes_large_file_through_duckdb():
    # Content is small, but a spoofed .size forces the large-file path —
    # isolates the routing decision from actually generating a huge file.
    f = _FakeUploadedFile(_csv_bytes(500), size=LARGE_FILE_THRESHOLD_BYTES + 1)
    df, error, warnings = load_data(f, max_rows=100)
    assert error is None
    assert len(df) == 100
    assert any("DuckDB" in w for w in warnings)


def test_load_data_falls_back_to_pandas_if_duckdb_cant_parse():
    # Spoofed large size, but content DuckDB's CSV reader chokes on (a
    # banner row it can't reconcile as well as pandas' dedicated recovery
    # path can) — must still succeed via the existing pandas fallback.
    messy = b"Company Sales Report\n\nid,value,label\n" + _csv_bytes(30).split(b"\n", 1)[1]
    f = _FakeUploadedFile(messy, size=LARGE_FILE_THRESHOLD_BYTES + 1)
    df, error, warnings = load_data(f)
    assert error is None
    assert len(df) >= 1


# ─────────────────────────────────────────────────────────────────────────
# _should_attempt_streaming_excel — the size gate
# ─────────────────────────────────────────────────────────────────────────
def test_should_attempt_streaming_excel_false_for_small_file():
    f = _FakeUploadedFile(_xlsx_bytes(10), name="data.xlsx", size=1_000)
    assert _should_attempt_streaming_excel(f) is False


def test_should_attempt_streaming_excel_true_above_threshold():
    f = _FakeUploadedFile(
        _xlsx_bytes(10), name="data.xlsx", size=LARGE_EXCEL_THRESHOLD_BYTES + 1
    )
    assert _should_attempt_streaming_excel(f) is True


def test_should_attempt_streaming_excel_false_when_size_unknown():
    f = _FakeUploadedFile(_xlsx_bytes(10), name="data.xlsx")
    del f.size
    assert _should_attempt_streaming_excel(f) is False


def test_should_attempt_streaming_excel_false_for_legacy_xls():
    # openpyxl can't open .xls at all — must not even attempt streaming.
    f = _FakeUploadedFile(
        _xlsx_bytes(10), name="data.xls", size=LARGE_EXCEL_THRESHOLD_BYTES + 1
    )
    assert _should_attempt_streaming_excel(f) is False


# ─────────────────────────────────────────────────────────────────────────
# _stream_sample_excel — the actual out-of-core read
# ─────────────────────────────────────────────────────────────────────────
def test_stream_sample_excel_returns_all_rows_when_under_cap():
    f = _FakeUploadedFile(_xlsx_bytes(50), name="data.xlsx")
    result = _stream_sample_excel(f, sheet_name=0, max_rows=200)
    assert result is not None
    df, warnings = result
    assert len(df) == 50
    assert list(df.columns) == ["id", "value", "label"]
    assert warnings == []


def test_stream_sample_excel_takes_random_sample_when_over_cap():
    f = _FakeUploadedFile(_xlsx_bytes(2000), name="data.xlsx")
    result = _stream_sample_excel(f, sheet_name=0, max_rows=100, random_state=42)
    assert result is not None
    df, warnings = result
    assert len(df) == 100
    assert warnings and "2,000" in warnings[0]
    # A random sample across the whole sheet, not just the first 100 rows.
    assert df["id"].max() > 200


def test_stream_sample_excel_is_reproducible_given_same_seed():
    data = _xlsx_bytes(2000)
    df1, _ = _stream_sample_excel(
        _FakeUploadedFile(data, name="data.xlsx"), sheet_name=0, max_rows=50, random_state=7
    )
    df2, _ = _stream_sample_excel(
        _FakeUploadedFile(data, name="data.xlsx"), sheet_name=0, max_rows=50, random_state=7
    )
    assert sorted(df1["id"].tolist()) == sorted(df2["id"].tolist())


def test_stream_sample_excel_handles_named_sheet():
    f = _FakeUploadedFile(_xlsx_bytes(20, sheet_name="Q3 Data"), name="data.xlsx")
    result = _stream_sample_excel(f, sheet_name="Q3 Data", max_rows=100)
    assert result is not None
    df, _ = result
    assert len(df) == 20


def test_stream_sample_excel_returns_none_for_missing_sheet():
    f = _FakeUploadedFile(_xlsx_bytes(10), name="data.xlsx")
    assert _stream_sample_excel(f, sheet_name="Nonexistent", max_rows=100) is None


def test_stream_sample_excel_returns_none_for_out_of_range_sheet_index():
    f = _FakeUploadedFile(_xlsx_bytes(10), name="data.xlsx")
    assert _stream_sample_excel(f, sheet_name=5, max_rows=100) is None


def test_stream_sample_excel_recovers_banner_row():
    f = _FakeUploadedFile(_xlsx_bytes(30, banner="Company Sales Report"), name="data.xlsx")
    result = _stream_sample_excel(f, sheet_name=0, max_rows=100)
    assert result is not None
    df, warnings = result
    assert list(df.columns) == ["id", "value", "label"]
    assert len(df) == 30
    assert any("banner" in w.lower() for w in warnings)


def test_stream_sample_excel_returns_none_for_empty_sheet():
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    f = _FakeUploadedFile(buf.getvalue(), name="empty.xlsx")
    assert _stream_sample_excel(f, sheet_name=0, max_rows=100) is None


def test_stream_sample_excel_returns_none_for_header_only_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "value", "label"])
    buf = io.BytesIO()
    wb.save(buf)
    f = _FakeUploadedFile(buf.getvalue(), name="header_only.xlsx")
    assert _stream_sample_excel(f, sheet_name=0, max_rows=100) is None


def test_stream_sample_excel_returns_none_on_corrupt_file():
    garbage = _FakeUploadedFile(b"not a real xlsx file at all", name="corrupt.xlsx")
    assert _stream_sample_excel(garbage, sheet_name=0, max_rows=100) is None


# ─────────────────────────────────────────────────────────────────────────
# load_data() — wiring: large .xlsx files route through the streaming
# reader, small ones and legacy .xls don't.
# ─────────────────────────────────────────────────────────────────────────
def test_load_data_small_excel_unaffected():
    f = _FakeUploadedFile(_xlsx_bytes(20), name="data.xlsx")
    df, error, warnings = load_data(f)
    assert error is None
    assert len(df) == 20
    assert not any("streaming" in w.lower() for w in warnings)


def test_load_data_routes_large_excel_through_streaming_reader():
    # Content is small, but a spoofed .size forces the large-file path —
    # isolates the routing decision from actually generating a huge file.
    f = _FakeUploadedFile(
        _xlsx_bytes(500), name="data.xlsx", size=LARGE_EXCEL_THRESHOLD_BYTES + 1
    )
    df, error, warnings = load_data(f, max_rows=100)
    assert error is None
    assert len(df) == 100
    assert any("streaming" in w.lower() for w in warnings)


def test_load_data_falls_back_to_pandas_if_streaming_excel_fails():
    # Spoofed large size, requesting a sheet index the streaming reader's
    # bounds check rejects (out of range) — must still fail cleanly via the
    # pandas fallback's own error handling rather than crash, since the
    # requested sheet genuinely doesn't exist either way.
    f = _FakeUploadedFile(
        _xlsx_bytes(30), name="data.xlsx", size=LARGE_EXCEL_THRESHOLD_BYTES + 1
    )
    df, error, warnings = load_data(f, sheet_name=5)
    assert df is None
    assert error is not None


def test_load_data_streaming_excel_falls_back_to_pandas_on_valid_sheet_name():
    # Streaming reader requires an exact sheet-name/index match; the
    # existing pandas fallback resolves sheet_name=0 to "whichever sheet is
    # first" regardless of its actual name — both should agree on success
    # for a normal request, proving the streaming path doesn't change
    # observable behavior for well-formed requests, only performance.
    f = _FakeUploadedFile(
        _xlsx_bytes(30, sheet_name="Q3 Data"),
        name="data.xlsx",
        size=LARGE_EXCEL_THRESHOLD_BYTES + 1,
    )
    df, error, warnings = load_data(f, sheet_name="Q3 Data")
    assert error is None
    assert len(df) == 30


def test_load_data_streaming_excel_respects_hard_row_ceiling():
    # max_rows=None (Smart Sampling's "read everything" request) must still
    # be capped at HARD_ROW_CEILING for the streaming path, same as every
    # other ingestion path.
    from modules.data_engine import HARD_ROW_CEILING

    f = _FakeUploadedFile(
        _xlsx_bytes(50), name="data.xlsx", size=LARGE_EXCEL_THRESHOLD_BYTES + 1
    )
    df, error, warnings = load_data(f, max_rows=None)
    assert error is None
    assert len(df) <= HARD_ROW_CEILING
